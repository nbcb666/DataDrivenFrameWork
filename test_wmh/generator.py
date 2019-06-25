# -*- coding: utf-8 -*-
# 1. load module
from openpyxl import *
dest_filename = 'test.xlsx'
# 2. load Workbook from existed file
wb=load_workbook(u'C:\\wmh\\DataDrivenFrameWork\\testData\\1.xlsx')
# 3. get a WorkSheet
ws=wb['Sheet1']
# 4. modify data
ws['A4']=100
ws.cell(row=1,column=2,value=100)
#5. read data
#read data
# for row in ws.rows: # 返回的row是一个tuple对象
#     for cell in row:
#         print 'row: %s  column: %s  value: %s' % (cell.row,cell.column,cell.value)
print type(ws)
RNum = 1
for row in ws.rows: # 返回的row是一个tuple对象
    for cell in row:
        #print 'row: %s  column: %s  value: %s' % (cell.row,cell.column,cell.value)
        if cell.column == 3:
            #print 'row: %s  column: %s  value: %s' % (cell.row,cell.column,cell.value)
            print 'value: %s' % (cell.value)


print '********************************'




b = ws.rows
print b
# 6. save Workbook to file
wb.save(dest_filename)